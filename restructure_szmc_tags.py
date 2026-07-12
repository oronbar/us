from __future__ import annotations

import argparse
import csv
import re
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl


DEFAULT_SOURCE_ROOT = Path(
    r"C:\Users\Oron\OneDrive - Technion\DS\Tags_SZMC\VVI\Anonymous"
)
DEFAULT_TARGET_ROOT = Path(r"C:\Users\Oron\OneDrive - Technion\DS\Tags_SZMC\VVI")
DEFAULT_REPORT = Path(r"C:\Users\Oron\OneDrive - Technion\DS\Report_SZMC_oron.xlsx")

CHAMBER_COLUMNS = ("2-Chambers", "3-Chambers", "4-Chambers")
DICOM_TOKEN_RE = re.compile(r"1\.2\.840\.[^\s,;]+?\.dcm", re.IGNORECASE)
INVALID_PATH_CHARS_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


@dataclass(frozen=True)
class MappingEntry:
    dicom_name: str
    chamber: str
    row_number: int
    short_id: str
    initials: str
    study_date: str
    patient_folder: str
    visit_folder: str

    @property
    def relative_destination(self) -> Path:
        return Path(self.patient_folder) / self.visit_folder / self.dicom_name


def clean_component(value: object, fallback: str) -> str:
    text = "" if value is None else str(value).strip()
    if not text:
        text = fallback
    text = INVALID_PATH_CHARS_RE.sub("_", text)
    text = re.sub(r"\s+", "_", text)
    text = text.strip(" ._")
    return text or fallback


def date_folder(value: object) -> str:
    text = clean_component(value, "UNKNOWN_DATE")
    match = re.match(r"^(\d{4})_(\d{2})_(\d{2})", text)
    if match:
        return "_".join(match.groups())
    return text


def load_mapping(report_path: Path) -> list[MappingEntry]:
    workbook = openpyxl.load_workbook(report_path, read_only=True, data_only=True)
    worksheet = workbook.active
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {value: index for index, value in enumerate(header_row) if value}

    required_columns = {"Short ID", "Name intials", "Study Date", *CHAMBER_COLUMNS}
    missing_columns = sorted(required_columns - set(headers))
    if missing_columns:
        raise ValueError(f"Missing required columns in workbook: {', '.join(missing_columns)}")

    entries: list[MappingEntry] = []
    seen_dicoms: dict[str, MappingEntry] = {}

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        short_id = clean_component(row[headers["Short ID"]], f"ROW_{row_number}")
        initials = clean_component(row[headers["Name intials"]], "UNKNOWN")
        visit = date_folder(row[headers["Study Date"]])
        patient = f"{short_id}_{initials}"

        for chamber in CHAMBER_COLUMNS:
            cell = row[headers[chamber]]
            if cell is None:
                continue

            for dicom_name in DICOM_TOKEN_RE.findall(str(cell)):
                entry = MappingEntry(
                    dicom_name=dicom_name,
                    chamber=chamber,
                    row_number=row_number,
                    short_id=short_id,
                    initials=initials,
                    study_date="" if row[headers["Study Date"]] is None else str(row[headers["Study Date"]]),
                    patient_folder=patient,
                    visit_folder=visit,
                )

                previous = seen_dicoms.get(dicom_name)
                if previous is not None and previous != entry:
                    raise ValueError(
                        "DICOM folder appears in multiple workbook rows: "
                        f"{dicom_name} at rows {previous.row_number} and {row_number}"
                    )

                seen_dicoms[dicom_name] = entry
                entries.append(entry)

    return entries


def write_csv(path: Path, fieldnames: Iterable[str], rows: Iterable[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fieldnames))
        writer.writeheader()
        writer.writerows(rows)


def directory_complete(source: Path, target: Path) -> bool:
    if not target.is_dir():
        return False

    for source_file in source.rglob("*"):
        if not source_file.is_file():
            continue

        relative_path = source_file.relative_to(source)
        target_file = target / relative_path
        if not target_file.is_file():
            return False
        if target_file.stat().st_size != source_file.stat().st_size:
            return False

    return True


def robocopy_directory(source: Path, target: Path) -> None:
    target.mkdir(parents=True, exist_ok=True)
    result = subprocess.run(
        [
            "robocopy",
            str(source),
            str(target),
            "/E",
            "/COPY:DAT",
            "/DCOPY:DAT",
            "/R:3",
            "/W:2",
            "/NP",
        ],
        capture_output=True,
    )

    if result.returncode >= 8:
        stdout = result.stdout.decode(errors="replace")
        stderr = result.stderr.decode(errors="replace")
        raise RuntimeError(
            f"robocopy failed with exit code {result.returncode}\n"
            f"STDOUT:\n{stdout}\nSTDERR:\n{stderr}"
        )


def copy_entry(
    entry: MappingEntry,
    source_root: Path,
    target_root: Path,
    dry_run: bool,
) -> str:
    source = source_root / entry.dicom_name
    target = target_root / entry.relative_destination

    if not source.is_dir():
        return "missing_source"

    target_exists = target.exists()
    if target_exists and directory_complete(source, target):
        return "exists"

    if dry_run:
        return "would_update_incomplete" if target_exists else "would_copy"

    robocopy_directory(source, target)
    if not directory_complete(source, target):
        raise RuntimeError(f"Copy verification failed for {target}")
    return "updated_incomplete" if target_exists else "copied"


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Copy SZMC Anonymous tag folders into patient/date/DICOM folders."
    )
    parser.add_argument("--source-root", type=Path, default=DEFAULT_SOURCE_ROOT)
    parser.add_argument("--target-root", type=Path, default=DEFAULT_TARGET_ROOT)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument(
        "--log-dir",
        type=Path,
        default=Path.cwd() / "szmc_restructure_logs",
        help="Directory for CSV audit logs.",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Actually copy folders. Without this flag, only a dry-run is performed.",
    )
    args = parser.parse_args()

    source_root = args.source_root
    target_root = args.target_root
    report = args.report
    dry_run = not args.apply

    if not source_root.is_dir():
        raise FileNotFoundError(f"Source root not found: {source_root}")
    if not report.is_file():
        raise FileNotFoundError(f"Workbook not found: {report}")

    entries = load_mapping(report)
    by_dicom = {entry.dicom_name: entry for entry in entries}
    source_dirs = {path.name for path in source_root.iterdir() if path.is_dir()}

    result_rows: list[dict[str, object]] = []
    status_counts: dict[str, int] = {}
    for entry in sorted(entries, key=lambda item: item.relative_destination.as_posix()):
        status = copy_entry(entry, source_root, target_root, dry_run)
        status_counts[status] = status_counts.get(status, 0) + 1
        result_rows.append(
            {
                "status": status,
                "source": str(source_root / entry.dicom_name),
                "destination": str(target_root / entry.relative_destination),
                "row_number": entry.row_number,
                "short_id": entry.short_id,
                "initials": entry.initials,
                "study_date": entry.study_date,
                "visit_folder": entry.visit_folder,
                "chamber": entry.chamber,
                "dicom_name": entry.dicom_name,
            }
        )

    unmapped_rows = [
        {"source": str(source_root / name), "dicom_name": name}
        for name in sorted(source_dirs - set(by_dicom))
    ]

    log_dir = args.log_dir
    write_csv(
        log_dir / ("dry_run_mapping.csv" if dry_run else "copy_mapping.csv"),
        (
            "status",
            "source",
            "destination",
            "row_number",
            "short_id",
            "initials",
            "study_date",
            "visit_folder",
            "chamber",
            "dicom_name",
        ),
        result_rows,
    )
    write_csv(log_dir / "unmapped_source_dirs.csv", ("source", "dicom_name"), unmapped_rows)

    print(f"Mode: {'dry-run' if dry_run else 'apply'}")
    print(f"Source folders: {len(source_dirs)}")
    print(f"Workbook DICOM mappings: {len(entries)}")
    print(f"Unmapped source folders: {len(unmapped_rows)}")
    for status, count in sorted(status_counts.items()):
        print(f"{status}: {count}")
    print(f"Logs: {log_dir}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
